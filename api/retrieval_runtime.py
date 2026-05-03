from __future__ import annotations

import io
import os
import sys
import zipfile
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import httpx
import openpyxl
import torch
import torch.nn.functional as F
from openpyxl.utils.exceptions import InvalidFileException
from transformers import AutoTokenizer

FETCH_TIMEOUT_SEC = 30
MAX_RESULTS_PER_EXCEL = 5


class ErrorCode(str, Enum):
    VALIDATION_ERROR = "VALIDATION_ERROR"
    JOB_NOT_FOUND = "JOB_NOT_FOUND"
    FILE_TOO_LARGE = "FILE_TOO_LARGE"
    UNSUPPORTED_FILE_TYPE = "UNSUPPORTED_FILE_TYPE"
    FILE_ENCRYPTED = "FILE_ENCRYPTED"
    FILE_CORRUPTED = "FILE_CORRUPTED"
    SOURCE_URL_INVALID = "SOURCE_URL_INVALID"
    SOURCE_URL_FETCH_FAILED = "SOURCE_URL_FETCH_FAILED"
    INTERNAL_ERROR = "INTERNAL_ERROR"


class RetrievalRuntimeError(Exception):
    def __init__(self, code: ErrorCode, message: str) -> None:
        super().__init__(message)
        self.code = code


@dataclass
class RetrievalResult:
    sheets: List[Dict[str, Any]]
    debug: Dict[str, Any]


class RetrievalRuntime:
    def __init__(
        self,
        repo_root: str,
        stage1_ckpt: str,
        stage2_ckpt: str,
        backbone_dir: str,
        tokenizer_dir: str,
        data_dir: str,
        device: Optional[str] = None,
        max_sheet_length: int = 256,
        max_query_length: int = 64,
        max_workspace_size: int = 10,
        max_header_texts: int = 12,
        top_k: int = MAX_RESULTS_PER_EXCEL,
        model_name_for_init: Optional[str] = None,
    ) -> None:
        self.repo_root = Path(repo_root)
        self.stage1_ckpt = Path(stage1_ckpt)
        self.stage2_ckpt = Path(stage2_ckpt)
        self.backbone_dir = Path(backbone_dir)
        self.tokenizer_dir = Path(tokenizer_dir)
        self.data_dir = Path(data_dir)
        self.device = torch.device(device or ("cuda" if torch.cuda.is_available() else "cpu"))
        self.max_sheet_length = max_sheet_length
        self.max_query_length = max_query_length
        self.max_workspace_size = max_workspace_size
        self.max_header_texts = max_header_texts
        self.top_k = top_k
        self.model_name_for_init = model_name_for_init or str(self.backbone_dir)

        self.tokenizer = None
        self.model = None
        self._loaded = False

    @classmethod
    def from_defaults(cls) -> "RetrievalRuntime":
        repo_root = "/root/sheetagentresearch/sheetagent_paper"
        return cls(
            repo_root=repo_root,
            stage1_ckpt=f"{repo_root}/best_model/classifier.pt",
            stage2_ckpt=f"{repo_root}/outputs/stage2_gtn_v2/stage2_gtn_v2_stable_lr15e5_ep50/best.pt",
            backbone_dir=f"{repo_root}/best_model/backbone",
            tokenizer_dir=f"{repo_root}/best_model",
            data_dir=f"{repo_root}/data",
        )

    def load(self) -> None:
        if self._loaded:
            return

        if str(self.repo_root) not in sys.path:
            sys.path.insert(0, str(self.repo_root))

        from stage2_gtn_v2 import Stage2GTNModelV2  # type: ignore

        self.tokenizer = AutoTokenizer.from_pretrained(str(self.tokenizer_dir), local_files_only=True)

        self.model = Stage2GTNModelV2(
            model_name=self.model_name_for_init,
            local_files_only=True,
            embedding_strategy="cls",
            use_layer_mix=False,
            use_extra_position_embedding=False,
            position_embedding_scale=1.0,
            max_length=self.max_sheet_length,
            normalize_embeddings=True,
            graph_dropout=0.1,
            num_gcn_layers=1,
            freeze_backbone=True,
            gtn_channels=4,
            gtn_layers=2,
        )

        if self.stage1_ckpt.exists():
            self.model.load_stage1_checkpoint(str(self.stage1_ckpt))

        stage2_obj = torch.load(str(self.stage2_ckpt), map_location="cpu")
        stage2_state = stage2_obj.get("state_dict", stage2_obj)
        self.model.load_state_dict(stage2_state, strict=False)

        self.model.to(self.device)
        self.model.eval()
        self._loaded = True

    async def fetch_excel(self, excel_url: str) -> bytes:
        if not excel_url.lower().startswith("https://"):
            raise RetrievalRuntimeError(ErrorCode.SOURCE_URL_INVALID, "excel_url 必须为 https://")

        try:
            async with httpx.AsyncClient(timeout=FETCH_TIMEOUT_SEC, follow_redirects=True) as client:
                resp = await client.get(excel_url)
                resp.raise_for_status()
        except httpx.HTTPError as exc:
            raise RetrievalRuntimeError(ErrorCode.SOURCE_URL_FETCH_FAILED, f"下载失败: {exc}") from exc

        content_type = (resp.headers.get("content-type") or "").lower()
        if not (
            excel_url.lower().endswith((".xlsx", ".xlsm"))
            or "spreadsheetml" in content_type
            or "excel" in content_type
            or "application/octet-stream" in content_type
        ):
            raise RetrievalRuntimeError(ErrorCode.UNSUPPORTED_FILE_TYPE, f"不支持的文件类型: {content_type or 'unknown'}")

        return resp.content

    async def retrieve_from_url(self, excel_url: str, query: str) -> RetrievalResult:
        content = await self.fetch_excel(excel_url)
        return self.retrieve_from_bytes(content, query)

    def retrieve_from_bytes(self, file_bytes: bytes, query: str) -> RetrievalResult:
        self.load()
        assert self.model is not None
        assert self.tokenizer is not None

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except InvalidFileException as exc:
            msg = str(exc).lower()
            if "encrypted" in msg or "password" in msg:
                raise RetrievalRuntimeError(ErrorCode.FILE_ENCRYPTED, "Excel 文件已加密，无法解析") from exc
            raise RetrievalRuntimeError(ErrorCode.UNSUPPORTED_FILE_TYPE, "文件类型不受支持") from exc
        except zipfile.BadZipFile as exc:
            raise RetrievalRuntimeError(ErrorCode.FILE_CORRUPTED, "Excel 文件损坏或内容非法") from exc
        except Exception as exc:  # noqa: BLE001
            msg = str(exc).lower()
            if "encrypted" in msg or "password" in msg:
                raise RetrievalRuntimeError(ErrorCode.FILE_ENCRYPTED, "Excel 文件已加密，无法解析") from exc
            raise RetrievalRuntimeError(ErrorCode.INTERNAL_ERROR, f"解析 Excel 失败: {exc}") from exc

        all_sheets = self._extract_workbook_sheets(workbook)
        if not all_sheets:
            return RetrievalResult(sheets=[], debug={"reason": "empty_workbook"})

        pre_scores = self._stage1_prerank(query=query, sheets=all_sheets)
        pre_scores.sort(key=lambda x: x["score"], reverse=True)
        workspace = pre_scores[: min(self.max_workspace_size, len(pre_scores))]

        stage2_scores = self._stage2_rerank(query=query, sheets=workspace)
        stage2_scores.sort(key=lambda x: x["score"], reverse=True)
        unique = []
        seen = set()
        for item in stage2_scores:
            key = (item["sheet_name"], item["sheet_index"])
            if key in seen:
                continue
            seen.add(key)
            unique.append({"sheet_name": item["sheet_name"], "sheet_index": item["sheet_index"]})
            if len(unique) >= self.top_k:
                break

        return RetrievalResult(
            sheets=unique,
            debug={"num_sheets_total": len(all_sheets), "num_sheets_workspace": len(workspace)},
        )

    def _extract_workbook_sheets(self, workbook: openpyxl.Workbook) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        for idx, ws in enumerate(workbook.worksheets):
            headers = self._extract_headers(ws)
            text = self._sheet_to_text(ws.title, int(ws.max_row or 0), int(ws.max_column or 0), headers)
            out.append(
                {
                    "sheet_name": ws.title,
                    "sheet_index": idx,
                    "num_rows": int(ws.max_row or 0),
                    "num_cols": int(ws.max_column or 0),
                    "headers": headers,
                    "text": text,
                }
            )
        return out

    def _extract_headers(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> List[str]:
        headers: List[str] = []
        max_cols = min(int(ws.max_column or 0), self.max_header_texts)
        for col_idx in range(1, max_cols + 1):
            value = ws.cell(row=1, column=col_idx).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                headers.append(text)
        return headers

    def _sheet_to_text(self, sheet_name: str, num_rows: int, num_cols: int, headers: Sequence[str]) -> str:
        parts = [f"name: {sheet_name}", f"shape: {num_rows} x {num_cols}"]
        if headers:
            parts.append("columns: " + " | ".join(headers[: self.max_header_texts]))
        return " ; ".join(parts)

    def _tokenize(self, text: str, max_length: int) -> Dict[str, torch.Tensor]:
        assert self.tokenizer is not None
        enc = self.tokenizer(
            text,
            max_length=max_length,
            padding="max_length",
            truncation=True,
            return_tensors="pt",
        )
        out = {
            "input_ids": enc["input_ids"].to(self.device),
            "attention_mask": enc["attention_mask"].to(self.device),
            "token_type_ids": enc.get("token_type_ids"),
        }
        if out["token_type_ids"] is not None:
            out["token_type_ids"] = out["token_type_ids"].to(self.device)
        return out

    @torch.no_grad()
    def _stage1_prerank(self, query: str, sheets: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        assert self.model is not None
        query_tok = self._tokenize(query, self.max_query_length)
        query_emb = self.model.encoder.encode(
            query_tok["input_ids"],
            query_tok["attention_mask"],
            query_tok["token_type_ids"],
        )

        out: List[Dict[str, Any]] = []
        for item in sheets:
            sheet_tok = self._tokenize(item["text"], self.max_sheet_length)
            sheet_emb = self.model.encoder.encode(
                sheet_tok["input_ids"],
                sheet_tok["attention_mask"],
                sheet_tok["token_type_ids"],
            )
            score = torch.sum(F.normalize(query_emb, dim=-1) * F.normalize(sheet_emb, dim=-1), dim=-1).item()
            out.append({**item, "score": float(score)})
        return out

    def _schema_prior(self, sheets: List[Dict[str, Any]]) -> torch.Tensor:
        n = len(sheets)
        mat = torch.zeros((1, n, n), dtype=torch.float32, device=self.device)
        for i in range(n):
            hi = {h.strip().lower() for h in sheets[i]["headers"] if str(h).strip()}
            for j in range(n):
                hj = {h.strip().lower() for h in sheets[j]["headers"] if str(h).strip()}
                inter = len(hi & hj)
                union = max(1, len(hi | hj))
                mat[0, i, j] = inter / union if (hi or hj) else 0.0
        return mat

    def _shape_prior(self, sheets: List[Dict[str, Any]]) -> torch.Tensor:
        n = len(sheets)
        mat = torch.zeros((1, n, n), dtype=torch.float32, device=self.device)
        for i in range(n):
            ri, ci = float(sheets[i]["num_rows"]), float(sheets[i]["num_cols"])
            for j in range(n):
                rj, cj = float(sheets[j]["num_rows"]), float(sheets[j]["num_cols"])
                row_sim = 1.0 - abs(ri - rj) / max(1.0, max(ri, rj))
                col_sim = 1.0 - abs(ci - cj) / max(1.0, max(ci, cj))
                mat[0, i, j] = max(0.0, (row_sim + col_sim) / 2.0)
        return mat

    @torch.no_grad()
    def _stage2_rerank(self, query: str, sheets: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        assert self.model is not None
        n = len(sheets)
        query_tok = self._tokenize(query, self.max_query_length)

        input_ids = []
        attn_masks = []
        token_type_ids = []
        for item in sheets:
            tok = self._tokenize(item["text"], self.max_sheet_length)
            input_ids.append(tok["input_ids"].squeeze(0))
            attn_masks.append(tok["attention_mask"].squeeze(0))
            if tok["token_type_ids"] is not None:
                token_type_ids.append(tok["token_type_ids"].squeeze(0))

        workspace_input_ids = torch.stack(input_ids, dim=0).unsqueeze(0).to(self.device)
        workspace_attention_mask = torch.stack(attn_masks, dim=0).unsqueeze(0).to(self.device)
        workspace_token_type_ids = None
        if len(token_type_ids) == len(sheets):
            workspace_token_type_ids = torch.stack(token_type_ids, dim=0).unsqueeze(0).to(self.device)

        node_mask = torch.ones((1, n), dtype=torch.float32, device=self.device)
        outputs = self.model(
            query_input_ids=query_tok["input_ids"],
            query_attention_mask=query_tok["attention_mask"],
            workspace_input_ids=workspace_input_ids,
            workspace_attention_mask=workspace_attention_mask,
            node_mask=node_mask,
            schema_prior=self._schema_prior(sheets),
            shape_prior=self._shape_prior(sheets),
            query_token_type_ids=query_tok["token_type_ids"],
            workspace_token_type_ids=workspace_token_type_ids,
        )

        probs = outputs["node_probs"][0].detach().cpu().tolist()
        cos = outputs["node_cos"][0].detach().cpu().tolist()
        reranked: List[Dict[str, Any]] = []
        for item, prob, cosine in zip(sheets, probs, cos):
            final_score = 0.7 * float(prob) + 0.3 * max(0.0, float(cosine))
            reranked.append(
                {
                    "sheet_name": item["sheet_name"],
                    "sheet_index": item["sheet_index"],
                    "score": final_score,
                }
            )
        return reranked
